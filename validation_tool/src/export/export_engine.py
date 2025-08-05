"""
Export Engine for Information Validation Tool

Provides comprehensive export capabilities including:
- PDF reports with charts and formatting
- Excel spreadsheets with multiple sheets
- CSV data exports
- Custom report templates
"""

import json
import csv
import io
import base64
from datetime import datetime
from typing import Dict, List, Any, Optional, Union
from reportlab.lib import colors
from reportlab.lib.pagesizes import letter, A4
from reportlab.platypus import SimpleDocTemplate, Table, TableStyle, Paragraph, Spacer, Image
from reportlab.lib.styles import getSampleStyleSheet, ParagraphStyle
from reportlab.lib.units import inch
from reportlab.graphics.shapes import Drawing
from reportlab.graphics.charts.piecharts import Pie
from reportlab.graphics.charts.barcharts import VerticalBarChart
import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.chart import PieChart, BarChart, Reference
import matplotlib.pyplot as plt
import matplotlib.patches as patches
from io import BytesIO

class ExportEngine:
    """Advanced export engine supporting multiple formats"""
    
    def __init__(self):
        self.styles = getSampleStyleSheet()
        self._setup_custom_styles()
    
    def _setup_custom_styles(self):
        """Setup custom styles for PDF generation"""
        # Title style
        self.title_style = ParagraphStyle(
            'CustomTitle',
            parent=self.styles['Heading1'],
            fontSize=24,
            spaceAfter=30,
            textColor=colors.HexColor('#2c3e50'),
            alignment=1  # Center alignment
        )
        
        # Heading style
        self.heading_style = ParagraphStyle(
            'CustomHeading',
            parent=self.styles['Heading2'],
            fontSize=16,
            spaceAfter=12,
            textColor=colors.HexColor('#34495e'),
            borderWidth=1,
            borderColor=colors.HexColor('#bdc3c7'),
            borderPadding=5
        )
        
        # Body style
        self.body_style = ParagraphStyle(
            'CustomBody',
            parent=self.styles['Normal'],
            fontSize=11,
            spaceAfter=6,
            textColor=colors.HexColor('#2c3e50')
        )
    
    def export_validation_results_pdf(self, validation_data: Dict[str, Any], 
                                    filename: str = None) -> bytes:
        """Export validation results as formatted PDF report"""
        
        if filename is None:
            filename = f"validation_report_{datetime.now().strftime('%Y%m%d_%H%M%S')}.pdf"
        
        # Create PDF buffer
        buffer = BytesIO()
        doc = SimpleDocTemplate(buffer, pagesize=A4, topMargin=1*inch)
        
        # Build story (content)
        story = []
        
        # Title
        title = Paragraph("Information Validation Report", self.title_style)
        story.append(title)
        story.append(Spacer(1, 20))
        
        # Executive Summary
        story.append(Paragraph("Executive Summary", self.heading_style))
        
        overall_score = validation_data.get('overall_score', 0)
        total_checks = validation_data.get('total_checks', 0)
        passed_checks = validation_data.get('passed_checks', 0)
        
        summary_text = f"""
        <b>Overall Score:</b> {overall_score:.1f}%<br/>
        <b>Total Checks:</b> {total_checks}<br/>
        <b>Passed Checks:</b> {passed_checks}<br/>
        <b>Generated:</b> {datetime.now().strftime('%Y-%m-%d %H:%M:%S')}
        """
        
        story.append(Paragraph(summary_text, self.body_style))
        story.append(Spacer(1, 20))
        
        # Score visualization
        if overall_score > 0:
            chart_image = self._create_score_chart(overall_score)
            if chart_image:
                story.append(chart_image)
                story.append(Spacer(1, 20))
        
        # Category Results
        if 'category_results' in validation_data:
            story.append(Paragraph("Category Results", self.heading_style))
            
            # Create category table
            category_data = [['Category', 'Score', 'Status', 'Issues']]
            
            for category, results in validation_data['category_results'].items():
                score = results.get('score', 0)
                status = results.get('status', 'unknown')
                issues_count = len(results.get('issues', []))
                
                category_data.append([
                    category.replace('_', ' ').title(),
                    f"{score:.1f}%",
                    status.title(),
                    str(issues_count)
                ])
            
            category_table = Table(category_data, colWidths=[2.5*inch, 1*inch, 1*inch, 1*inch])
            category_table.setStyle(TableStyle([
                ('BACKGROUND', (0, 0), (-1, 0), colors.HexColor('#3498db')),
                ('TEXTCOLOR', (0, 0), (-1, 0), colors.whitesmoke),
                ('ALIGN', (0, 0), (-1, -1), 'CENTER'),
                ('FONTNAME', (0, 0), (-1, 0), 'Helvetica-Bold'),
                ('FONTSIZE', (0, 0), (-1, 0), 12),
                ('BOTTOMPADDING', (0, 0), (-1, 0), 12),
                ('BACKGROUND', (0, 1), (-1, -1), colors.beige),
                ('GRID', (0, 0), (-1, -1), 1, colors.black)
            ]))
            
            story.append(category_table)
            story.append(Spacer(1, 20))
        
        # Detailed Issues - fix field name mismatch
        issues_data = validation_data.get('detailed_issues') or validation_data.get('issues', [])
        if issues_data:
            story.append(Paragraph("Detailed Issues", self.heading_style))
            
            for i, issue in enumerate(issues_data[:10], 1):  # Limit to top 10
                issue_text = f"""
                <b>{i}. {issue.get('title', issue.get('description', 'Unknown Issue'))}</b><br/>
                <b>Category:</b> {issue.get('category', 'N/A')}<br/>
                <b>Severity:</b> {issue.get('severity', 'Medium')}<br/>
                <b>Description:</b> {issue.get('description', 'No description available')}<br/>
                """
                
                if 'recommendation' in issue:
                    issue_text += f"<b>Recommendation:</b> {issue['recommendation']}<br/>"
                
                story.append(Paragraph(issue_text, self.body_style))
                story.append(Spacer(1, 10))
        
        # Recommendations
        if 'recommendations' in validation_data and validation_data['recommendations']:
            story.append(Paragraph("Recommendations", self.heading_style))
            
            for i, rec in enumerate(validation_data['recommendations'][:5], 1):
                rec_text = f"""
                <b>{i}. {rec.get('title', 'Recommendation')}</b><br/>
                <b>Priority:</b> {rec.get('priority', 'Medium').title()}<br/>
                {rec.get('description', 'No description available')}
                """
                
                story.append(Paragraph(rec_text, self.body_style))
                story.append(Spacer(1, 10))
        
        # Build PDF
        doc.build(story)
        
        # Get PDF bytes
        buffer.seek(0)
        pdf_bytes = buffer.getvalue()
        buffer.close()
        
        return pdf_bytes
    
    def export_validation_results_excel(self, validation_data: Dict[str, Any], 
                                      filename: str = None) -> bytes:
        """Export validation results as Excel workbook with multiple sheets"""
        
        if filename is None:
            filename = f"validation_results_{datetime.now().strftime('%Y%m%d_%H%M%S')}.xlsx"
        
        # Create workbook
        wb = openpyxl.Workbook()
        
        # Remove default sheet
        wb.remove(wb.active)
        
        # Summary sheet
        self._create_summary_sheet(wb, validation_data)
        
        # Category details sheet
        if 'category_results' in validation_data:
            self._create_category_sheet(wb, validation_data['category_results'])
        
        # Issues sheet - fix field name mismatch
        if 'detailed_issues' in validation_data:
            self._create_issues_sheet(wb, validation_data['detailed_issues'])
        elif 'issues' in validation_data:  # fallback for backward compatibility
            self._create_issues_sheet(wb, validation_data['issues'])
        
        # Recommendations sheet
        if 'recommendations' in validation_data:
            self._create_recommendations_sheet(wb, validation_data['recommendations'])
        
        # Raw data sheet
        self._create_raw_data_sheet(wb, validation_data)
        
        # Save to buffer
        buffer = BytesIO()
        wb.save(buffer)
        buffer.seek(0)
        
        excel_bytes = buffer.getvalue()
        buffer.close()
        
        return excel_bytes
    
    def export_validation_results_csv(self, validation_data: Dict[str, Any], 
                                    export_type: str = 'summary') -> str:
        """Export validation results as CSV data"""
        
        output = io.StringIO()
        
        if export_type == 'summary':
            writer = csv.writer(output)
            
            # Header
            writer.writerow(['Metric', 'Value'])
            
            # Summary data
            writer.writerow(['Overall Score', f"{validation_data.get('overall_score', 0):.1f}%"])
            writer.writerow(['Total Checks', validation_data.get('total_checks', 0)])
            writer.writerow(['Passed Checks', validation_data.get('passed_checks', 0)])
            writer.writerow(['Failed Checks', validation_data.get('failed_checks', 0)])
            writer.writerow(['Generated At', datetime.now().strftime('%Y-%m-%d %H:%M:%S')])
            
        elif export_type == 'categories':
            writer = csv.writer(output)
            
            # Header
            writer.writerow(['Category', 'Score', 'Status', 'Issues Count', 'Pass Rate'])
            
            # Category data
            if 'category_results' in validation_data:
                for category, results in validation_data['category_results'].items():
                    writer.writerow([
                        category.replace('_', ' ').title(),
                        f"{results.get('score', 0):.1f}",
                        results.get('status', 'unknown'),
                        len(results.get('issues', [])),
                        f"{results.get('pass_rate', 0):.1f}%"
                    ])
        
        elif export_type == 'issues':
            writer = csv.writer(output)
            
            # Header
            writer.writerow(['Issue ID', 'Title', 'Category', 'Severity', 'Description', 'Recommendation'])
            
            # Issues data - fix field name mismatch
            issues_data = validation_data.get('detailed_issues') or validation_data.get('issues', [])
            for i, issue in enumerate(issues_data, 1):
                writer.writerow([
                    i,
                    issue.get('title', issue.get('description', 'Unknown Issue')),  # fallback to description if no title
                    issue.get('category', 'N/A'),
                    issue.get('severity', 'Medium'),
                    issue.get('description', 'No description'),
                    issue.get('recommendation', 'No recommendation')
                ])
        
        csv_content = output.getvalue()
        output.close()
        
        return csv_content
    
    def _create_score_chart(self, score: float) -> Optional[Image]:
        """Create a score visualization chart for PDF"""
        try:
            # Create matplotlib figure
            fig, ax = plt.subplots(figsize=(6, 4))
            
            # Create gauge chart
            theta = (score / 100) * 180  # Convert to degrees (0-180)
            
            # Draw gauge background
            gauge_bg = patches.Wedge((0.5, 0), 0.4, 0, 180, 
                                   facecolor='lightgray', edgecolor='black')
            ax.add_patch(gauge_bg)
            
            # Draw score wedge
            if score >= 80:
                color = 'green'
            elif score >= 60:
                color = 'orange'
            else:
                color = 'red'
            
            score_wedge = patches.Wedge((0.5, 0), 0.4, 0, theta, 
                                      facecolor=color, alpha=0.7)
            ax.add_patch(score_wedge)
            
            # Add score text
            ax.text(0.5, 0.2, f'{score:.1f}%', ha='center', va='center', 
                   fontsize=20, fontweight='bold')
            
            ax.set_xlim(0, 1)
            ax.set_ylim(0, 0.6)
            ax.set_aspect('equal')
            ax.axis('off')
            ax.set_title('Overall Validation Score', fontsize=14, fontweight='bold')
            
            # Save to buffer
            img_buffer = BytesIO()
            plt.savefig(img_buffer, format='png', bbox_inches='tight', dpi=150)
            plt.close()
            
            img_buffer.seek(0)
            
            # Create ReportLab Image
            img = Image(img_buffer, width=4*inch, height=2.5*inch)
            return img
            
        except Exception as e:
            print(f"Error creating chart: {e}")
            return None
    
    def _create_summary_sheet(self, wb: openpyxl.Workbook, validation_data: Dict[str, Any]):
        """Create summary sheet in Excel workbook"""
        ws = wb.create_sheet("Summary")
        
        # Title
        ws['A1'] = "Validation Results Summary"
        ws['A1'].font = Font(size=16, bold=True)
        ws.merge_cells('A1:B1')
        
        # Summary data
        row = 3
        summary_items = [
            ('Overall Score', f"{validation_data.get('overall_score', 0):.1f}%"),
            ('Total Checks', validation_data.get('total_checks', 0)),
            ('Passed Checks', validation_data.get('passed_checks', 0)),
            ('Failed Checks', validation_data.get('failed_checks', 0)),
            ('Warnings', validation_data.get('warnings', 0)),
            ('Generated At', datetime.now().strftime('%Y-%m-%d %H:%M:%S'))
        ]
        
        for label, value in summary_items:
            ws[f'A{row}'] = label
            ws[f'B{row}'] = value
            ws[f'A{row}'].font = Font(bold=True)
            row += 1
        
        # Auto-adjust column widths
        for column in ws.columns:
            max_length = 0
            column_letter = None
            for cell in column:
                try:
                    # Skip merged cells
                    if hasattr(cell, 'column_letter'):
                        if column_letter is None:
                            column_letter = cell.column_letter
                        if len(str(cell.value)) > max_length:
                            max_length = len(str(cell.value))
                except:
                    pass
            
            if column_letter:
                adjusted_width = min(max_length + 2, 50)
                ws.column_dimensions[column_letter].width = adjusted_width
    
    def _create_category_sheet(self, wb: openpyxl.Workbook, category_results: Dict[str, Any]):
        """Create category results sheet in Excel workbook"""
        ws = wb.create_sheet("Category Results")
        
        # Headers
        headers = ['Category', 'Score', 'Status', 'Issues Count', 'Pass Rate']
        for col, header in enumerate(headers, 1):
            cell = ws.cell(row=1, column=col, value=header)
            cell.font = Font(bold=True)
            cell.fill = PatternFill(start_color="366092", end_color="366092", fill_type="solid")
            cell.font = Font(color="FFFFFF", bold=True)
        
        # Data
        row = 2
        for category, results in category_results.items():
            ws.cell(row=row, column=1, value=category.replace('_', ' ').title())
            ws.cell(row=row, column=2, value=f"{results.get('score', 0):.1f}%")
            ws.cell(row=row, column=3, value=results.get('status', 'unknown').title())
            ws.cell(row=row, column=4, value=len(results.get('issues', [])))
            ws.cell(row=row, column=5, value=f"{results.get('pass_rate', 0):.1f}%")
            row += 1
        
        # Auto-adjust column widths
        for column in ws.columns:
            max_length = 0
            column_letter = None
            for cell in column:
                try:
                    # Skip merged cells
                    if hasattr(cell, 'column_letter'):
                        if column_letter is None:
                            column_letter = cell.column_letter
                        if len(str(cell.value)) > max_length:
                            max_length = len(str(cell.value))
                except:
                    pass
            
            if column_letter:
                adjusted_width = min(max_length + 2, 50)
                ws.column_dimensions[column_letter].width = adjusted_width
    
    def _create_issues_sheet(self, wb: openpyxl.Workbook, issues: List[Dict[str, Any]]):
        """Create issues sheet in Excel workbook"""
        ws = wb.create_sheet("Issues")
        
        # Headers
        headers = ['ID', 'Title', 'Category', 'Severity', 'Description', 'Recommendation']
        for col, header in enumerate(headers, 1):
            cell = ws.cell(row=1, column=col, value=header)
            cell.font = Font(bold=True)
            cell.fill = PatternFill(start_color="E74C3C", end_color="E74C3C", fill_type="solid")
            cell.font = Font(color="FFFFFF", bold=True)
        
        # Data
        for row, issue in enumerate(issues, 2):
            ws.cell(row=row, column=1, value=row-1)
            ws.cell(row=row, column=2, value=issue.get('title', 'Unknown Issue'))
            ws.cell(row=row, column=3, value=issue.get('category', 'N/A'))
            ws.cell(row=row, column=4, value=issue.get('severity', 'Medium'))
            ws.cell(row=row, column=5, value=issue.get('description', 'No description'))
            ws.cell(row=row, column=6, value=issue.get('recommendation', 'No recommendation'))
        
        # Auto-adjust column widths
        for column in ws.columns:
            max_length = 0
            column_letter = None
            for cell in column:
                try:
                    # Skip merged cells
                    if hasattr(cell, 'column_letter'):
                        if column_letter is None:
                            column_letter = cell.column_letter
                        if len(str(cell.value)) > max_length:
                            max_length = len(str(cell.value))
                except:
                    pass
            
            if column_letter:
                adjusted_width = min(max_length + 2, 50)
                ws.column_dimensions[column_letter].width = adjusted_width
    
    def _create_recommendations_sheet(self, wb: openpyxl.Workbook, recommendations: List[Dict[str, Any]]):
        """Create recommendations sheet in Excel workbook"""
        ws = wb.create_sheet("Recommendations")
        
        # Headers
        headers = ['Priority', 'Title', 'Type', 'Description', 'Action']
        for col, header in enumerate(headers, 1):
            cell = ws.cell(row=1, column=col, value=header)
            cell.font = Font(bold=True)
            cell.fill = PatternFill(start_color="27AE60", end_color="27AE60", fill_type="solid")
            cell.font = Font(color="FFFFFF", bold=True)
        
        # Data
        for row, rec in enumerate(recommendations, 2):
            ws.cell(row=row, column=1, value=rec.get('priority', 'Medium').title())
            ws.cell(row=row, column=2, value=rec.get('title', 'Recommendation'))
            ws.cell(row=row, column=3, value=rec.get('type', 'general'))
            ws.cell(row=row, column=4, value=rec.get('description', 'No description'))
            ws.cell(row=row, column=5, value=rec.get('action', 'No action specified'))
        
        # Auto-adjust column widths
        for column in ws.columns:
            max_length = 0
            column_letter = None
            for cell in column:
                try:
                    # Skip merged cells
                    if hasattr(cell, 'column_letter'):
                        if column_letter is None:
                            column_letter = cell.column_letter
                        if len(str(cell.value)) > max_length:
                            max_length = len(str(cell.value))
                except:
                    pass
            
            if column_letter:
                adjusted_width = min(max_length + 2, 50)
                ws.column_dimensions[column_letter].width = adjusted_width
    
    def _create_raw_data_sheet(self, wb: openpyxl.Workbook, validation_data: Dict[str, Any]):
        """Create raw data sheet with complete validation data"""
        ws = wb.create_sheet("Raw Data")
        
        # Convert validation data to JSON string for display
        ws['A1'] = "Complete Validation Data (JSON)"
        ws['A1'].font = Font(size=14, bold=True)
        
        json_data = json.dumps(validation_data, indent=2)
        ws['A3'] = json_data
        
        # Adjust column width
        ws.column_dimensions['A'].width = 100
    
    def create_trends_report_pdf(self, trends_data: Dict[str, Any]) -> bytes:
        """Create comprehensive trends analysis PDF report"""
        buffer = BytesIO()
        doc = SimpleDocTemplate(buffer, pagesize=A4, topMargin=1*inch)
        
        story = []
        
        # Title
        title = Paragraph("Validation Trends Analysis Report", self.title_style)
        story.append(title)
        story.append(Spacer(1, 20))
        
        # Executive Summary
        story.append(Paragraph("Executive Summary", self.heading_style))
        
        overview = trends_data.get('overview', {})
        summary_text = f"""
        <b>Analysis Period:</b> {trends_data.get('period_days', 30)} days<br/>
        <b>Total Validations:</b> {overview.get('total_validations', 0)}<br/>
        <b>Average Score:</b> {overview.get('average_score', 0):.1f}%<br/>
        <b>Score Trend:</b> {overview.get('score_trend', 'stable').title()}<br/>
        <b>Generated:</b> {datetime.now().strftime('%Y-%m-%d %H:%M:%S')}
        """
        
        story.append(Paragraph(summary_text, self.body_style))
        story.append(Spacer(1, 20))
        
        # Category Trends
        if 'category_trends' in trends_data:
            story.append(Paragraph("Category Performance Trends", self.heading_style))
            
            category_data = [['Category', 'Average Score', 'Pass Rate', 'Trend']]
            
            for category, data in trends_data['category_trends'].items():
                category_data.append([
                    category.replace('_', ' ').title(),
                    f"{data.get('average_score', 0):.1f}%",
                    f"{data.get('pass_rate', 0):.1f}%",
                    data.get('score_trend', 'stable').title()
                ])
            
            category_table = Table(category_data, colWidths=[2*inch, 1.2*inch, 1.2*inch, 1.2*inch])
            category_table.setStyle(TableStyle([
                ('BACKGROUND', (0, 0), (-1, 0), colors.HexColor('#3498db')),
                ('TEXTCOLOR', (0, 0), (-1, 0), colors.whitesmoke),
                ('ALIGN', (0, 0), (-1, -1), 'CENTER'),
                ('FONTNAME', (0, 0), (-1, 0), 'Helvetica-Bold'),
                ('FONTSIZE', (0, 0), (-1, 0), 11),
                ('BOTTOMPADDING', (0, 0), (-1, 0), 12),
                ('BACKGROUND', (0, 1), (-1, -1), colors.beige),
                ('GRID', (0, 0), (-1, -1), 1, colors.black)
            ]))
            
            story.append(category_table)
            story.append(Spacer(1, 20))
        
        # Recommendations
        if 'recommendations' in trends_data:
            story.append(Paragraph("Key Recommendations", self.heading_style))
            
            for i, rec in enumerate(trends_data['recommendations'][:5], 1):
                rec_text = f"""
                <b>{i}. {rec.get('title', 'Recommendation')}</b><br/>
                <b>Priority:</b> {rec.get('priority', 'Medium').title()}<br/>
                {rec.get('description', 'No description available')}
                """
                
                story.append(Paragraph(rec_text, self.body_style))
                story.append(Spacer(1, 10))
        
        # Build PDF
        doc.build(story)
        
        buffer.seek(0)
        pdf_bytes = buffer.getvalue()
        buffer.close()
        
        return pdf_bytes

