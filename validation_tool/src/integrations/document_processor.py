"""
Document Processor for Site Survey and Install Plan Documents
Handles extraction and processing of content from various document formats
"""

import json
import re
import logging
from typing import Dict, List, Any, Optional
from dataclasses import dataclass
from io import BytesIO
import PyPDF2
from openpyxl import load_workbook

logger = logging.getLogger(__name__)

@dataclass
class ProcessedDocument:
    """Processed document with extracted content"""
    document_type: str
    source_url: str
    content: Dict[str, Any]
    metadata: Dict[str, Any]
    processing_errors: List[str]

class DocumentProcessor:
    """Process various document types for validation"""
    
    def __init__(self):
        self.logger = logging.getLogger(__name__)
        self.processors = {
            'site_survey_part1': self._process_site_survey_part1,
            'site_survey_part2': self._process_site_survey_part2,
            'install_plan_pdf': self._process_install_plan_pdf
        }
    
    def process_document(
        self, 
        document_type: str, 
        source_url: str, 
        content_data: bytes
    ) -> ProcessedDocument:
        """
        Process a document and extract structured content
        
        Args:
            document_type: Type of document (site_survey_part1, site_survey_part2, install_plan_pdf)
            source_url: Source URL or identifier
            content_data: Raw document content as bytes
            
        Returns:
            ProcessedDocument with extracted content
        """
        processor = self.processors.get(document_type)
        if not processor:
            return ProcessedDocument(
                document_type=document_type,
                source_url=source_url,
                content={},
                metadata={},
                processing_errors=[f"Unknown document type: {document_type}"]
            )
        
        try:
            return processor(source_url, content_data)
        except Exception as e:
            self.logger.error(f"Error processing {document_type}: {str(e)}")
            return ProcessedDocument(
                document_type=document_type,
                source_url=source_url,
                content={},
                metadata={'error': str(e)},
                processing_errors=[str(e)]
            )
    
    def _process_site_survey_part1(self, source_url: str, content_data: bytes) -> ProcessedDocument:
        """Process Site Survey Part 1 Excel document"""
        errors = []
        content = {}
        metadata = {}
        
        try:
            # Load Excel workbook
            workbook = load_workbook(BytesIO(content_data), data_only=True)
            metadata['worksheets'] = workbook.sheetnames
            
            # Process each expected worksheet
            expected_worksheets = [
                "Release Notes",
                "Project Details", 
                "VAST Cluster",
                "Rack Diagram",
                "Racks and Power",
                "Network Details",
                "VAST Hardware Details"
            ]
            
            for worksheet_name in expected_worksheets:
                if worksheet_name in workbook.sheetnames:
                    try:
                        worksheet_content = self._extract_worksheet_content(
                            workbook[worksheet_name], worksheet_name
                        )
                        content[worksheet_name] = worksheet_content
                    except Exception as e:
                        errors.append(f"Error processing worksheet {worksheet_name}: {str(e)}")
                        content[worksheet_name] = {'error': str(e)}
                else:
                    errors.append(f"Missing expected worksheet: {worksheet_name}")
            
            # Extract key metadata
            metadata.update(self._extract_site_survey_metadata(content))
            
        except Exception as e:
            errors.append(f"Error loading Excel workbook: {str(e)}")
        
        return ProcessedDocument(
            document_type='site_survey_part1',
            source_url=source_url,
            content=content,
            metadata=metadata,
            processing_errors=errors
        )
    
    def _process_site_survey_part2(self, source_url: str, content_data: bytes) -> ProcessedDocument:
        """Process Site Survey Part 2 Excel document"""
        errors = []
        content = {}
        metadata = {}
        
        try:
            # Load Excel workbook
            workbook = load_workbook(BytesIO(content_data), data_only=True)
            metadata['worksheets'] = workbook.sheetnames
            
            # Process each expected worksheet
            expected_worksheets = [
                "Release Notes",
                "Import from Part 1",
                "Customer Support",
                "Features and Configuration",
                "#2 IP Addresses",
                "Network Diagram"
            ]
            
            for worksheet_name in expected_worksheets:
                if worksheet_name in workbook.sheetnames:
                    try:
                        worksheet_content = self._extract_worksheet_content(
                            workbook[worksheet_name], worksheet_name
                        )
                        content[worksheet_name] = worksheet_content
                    except Exception as e:
                        errors.append(f"Error processing worksheet {worksheet_name}: {str(e)}")
                        content[worksheet_name] = {'error': str(e)}
                else:
                    errors.append(f"Missing expected worksheet: {worksheet_name}")
            
            # Extract key metadata
            metadata.update(self._extract_site_survey_metadata(content))
            
        except Exception as e:
            errors.append(f"Error loading Excel workbook: {str(e)}")
        
        return ProcessedDocument(
            document_type='site_survey_part2',
            source_url=source_url,
            content=content,
            metadata=metadata,
            processing_errors=errors
        )
    
    def _process_install_plan_pdf(self, source_url: str, content_data: bytes) -> ProcessedDocument:
        """Process Install Plan PDF document"""
        errors = []
        content = {}
        metadata = {}
        
        try:
            # Extract text from PDF
            pdf_reader = PyPDF2.PdfReader(BytesIO(content_data))
            metadata['total_pages'] = len(pdf_reader.pages)
            
            # Extract text from all pages
            full_text = ""
            page_contents = {}
            
            for page_num, page in enumerate(pdf_reader.pages):
                try:
                    page_text = page.extract_text()
                    page_contents[f"page_{page_num + 1}"] = page_text
                    full_text += page_text + "\n"
                except Exception as e:
                    errors.append(f"Error extracting text from page {page_num + 1}: {str(e)}")
            
            content['full_text'] = full_text
            content['pages'] = page_contents
            
            # Extract structured sections
            sections = self._extract_pdf_sections(full_text)
            content['sections'] = sections
            
            # Extract key information
            metadata.update(self._extract_install_plan_metadata(full_text, sections))
            
        except Exception as e:
            errors.append(f"Error processing PDF: {str(e)}")
        
        return ProcessedDocument(
            document_type='install_plan_pdf',
            source_url=source_url,
            content=content,
            metadata=metadata,
            processing_errors=errors
        )
    
    def _extract_worksheet_content(self, worksheet, worksheet_name: str) -> Dict[str, Any]:
        """Extract content from Excel worksheet"""
        content = {
            'name': worksheet_name,
            'max_row': worksheet.max_row,
            'max_column': worksheet.max_column,
            'cells': {},
            'tables': [],
            'key_values': {}
        }
        
        # Extract all cell values
        for row in range(1, min(worksheet.max_row + 1, 101)):  # Limit to first 100 rows
            for col in range(1, min(worksheet.max_column + 1, 27)):  # Limit to first 26 columns
                cell = worksheet.cell(row=row, column=col)
                if cell.value is not None:
                    cell_ref = f"{chr(64 + col)}{row}"  # Convert to A1 notation
                    content['cells'][cell_ref] = cell.value
        
        # Extract specific patterns based on worksheet type
        if worksheet_name == "Project Details":
            content['key_values'] = self._extract_project_details(worksheet)
        elif worksheet_name == "VAST Cluster":
            content['cluster_info'] = self._extract_cluster_info(worksheet)
        elif worksheet_name == "VAST Hardware Details":
            content['hardware_details'] = self._extract_hardware_details(worksheet)
        elif worksheet_name == "#2 IP Addresses":
            content['ip_addresses'] = self._extract_ip_addresses(worksheet)
        elif worksheet_name == "Release Notes":
            content['release_info'] = self._extract_release_notes(worksheet)
        
        return content
    
    def _extract_project_details(self, worksheet) -> Dict[str, Any]:
        """Extract project details from Project Details worksheet"""
        details = {}
        
        # Common project detail locations
        detail_mappings = {
            'customer_name': 'B17',
            'project_name': 'B18',
            'opportunity_id': 'B21',
            'project_manager': 'B22',
            'sales_engineer': 'B23'
        }
        
        for key, cell_ref in detail_mappings.items():
            try:
                col = ord(cell_ref[0]) - 64
                row = int(cell_ref[1:])
                cell_value = worksheet.cell(row=row, column=col).value
                if cell_value:
                    details[key] = str(cell_value).strip()
            except Exception as e:
                logger.warning(f"Error extracting {key} from {cell_ref}: {str(e)}")
        
        return details
    
    def _extract_cluster_info(self, worksheet) -> Dict[str, Any]:
        """Extract cluster information from VAST Cluster worksheet"""
        cluster_info = {}
        
        # Extract cluster configuration details
        try:
            # Look for cluster size information
            for row in range(1, min(worksheet.max_row + 1, 51)):
                for col in range(1, min(worksheet.max_column + 1, 11)):
                    cell_value = worksheet.cell(row=row, column=col).value
                    if cell_value and isinstance(cell_value, str):
                        if 'cluster' in cell_value.lower():
                            cluster_info['cluster_description'] = cell_value
                        elif 'node' in cell_value.lower():
                            cluster_info['node_info'] = cell_value
        except Exception as e:
            logger.warning(f"Error extracting cluster info: {str(e)}")
        
        return cluster_info
    
    def _extract_hardware_details(self, worksheet) -> Dict[str, Any]:
        """Extract hardware details from VAST Hardware Details worksheet"""
        hardware = {}
        
        try:
            # Extract hardware table
            hardware_items = []
            
            # Look for table headers and data
            for row in range(1, min(worksheet.max_row + 1, 101)):
                row_data = {}
                has_data = False
                
                for col in range(1, min(worksheet.max_column + 1, 11)):
                    cell_value = worksheet.cell(row=row, column=col).value
                    if cell_value:
                        col_letter = chr(64 + col)
                        row_data[col_letter] = cell_value
                        has_data = True
                
                if has_data:
                    hardware_items.append(row_data)
            
            hardware['items'] = hardware_items
            
        except Exception as e:
            logger.warning(f"Error extracting hardware details: {str(e)}")
        
        return hardware
    
    def _extract_ip_addresses(self, worksheet) -> Dict[str, Any]:
        """Extract IP address information from #2 IP Addresses worksheet"""
        ip_info = {}
        
        try:
            # Extract IP address table
            ip_entries = []
            
            for row in range(1, min(worksheet.max_row + 1, 101)):
                row_data = {}
                has_data = False
                
                for col in range(1, min(worksheet.max_column + 1, 11)):
                    cell_value = worksheet.cell(row=row, column=col).value
                    if cell_value:
                        col_letter = chr(64 + col)
                        row_data[col_letter] = cell_value
                        has_data = True
                
                if has_data:
                    ip_entries.append(row_data)
            
            ip_info['entries'] = ip_entries
            
            # Extract specific IP ranges and VLANs
            ip_ranges = []
            vlans = []
            
            for entry in ip_entries:
                for key, value in entry.items():
                    if value and isinstance(value, str):
                        # Look for IP address patterns
                        ip_pattern = r'\b(?:[0-9]{1,3}\.){3}[0-9]{1,3}\b'
                        if re.search(ip_pattern, value):
                            ip_ranges.append(value)
                        
                        # Look for VLAN patterns
                        if 'vlan' in value.lower() or re.search(r'\bv\d+\b', value.lower()):
                            vlans.append(value)
            
            ip_info['ip_ranges'] = ip_ranges
            ip_info['vlans'] = vlans
            
        except Exception as e:
            logger.warning(f"Error extracting IP addresses: {str(e)}")
        
        return ip_info
    
    def _extract_release_notes(self, worksheet) -> Dict[str, Any]:
        """Extract release notes information"""
        release_info = {}
        
        try:
            # Look for version information
            for row in range(1, min(worksheet.max_row + 1, 21)):
                for col in range(1, min(worksheet.max_column + 1, 6)):
                    cell_value = worksheet.cell(row=row, column=col).value
                    if cell_value and isinstance(cell_value, str):
                        if 'version' in cell_value.lower() or 'v' in cell_value.lower():
                            # Check adjacent cells for version number
                            for check_col in range(max(1, col-1), min(worksheet.max_column + 1, col+3)):
                                check_value = worksheet.cell(row=row, column=check_col).value
                                if check_value and re.search(r'v?\d+\.\d+\.\d+', str(check_value)):
                                    release_info['version'] = str(check_value)
                                    break
            
            # Look for approval information
            approvals = []
            for row in range(20, min(worksheet.max_row + 1, 51)):
                row_data = {}
                has_approval_data = False
                
                for col in range(1, min(worksheet.max_column + 1, 11)):
                    cell_value = worksheet.cell(row=row, column=col).value
                    if cell_value:
                        col_letter = chr(64 + col)
                        row_data[col_letter] = cell_value
                        if any(keyword in str(cell_value).lower() for keyword in ['approve', 'review', 'sign']):
                            has_approval_data = True
                
                if has_approval_data:
                    approvals.append(row_data)
            
            release_info['approvals'] = approvals
            
        except Exception as e:
            logger.warning(f"Error extracting release notes: {str(e)}")
        
        return release_info
    
    def _extract_pdf_sections(self, full_text: str) -> Dict[str, str]:
        """Extract sections from PDF text"""
        sections = {}
        
        # Common section patterns in install plans
        section_patterns = [
            (r'Prerequisites.*?(?=\n[A-Z][A-Za-z\s]+:|\nHardware|\nNetwork|\Z)', 'Prerequisites'),
            (r'Hardware.*?(?=\n[A-Z][A-Za-z\s]+:|\nNetwork|\nInstallation|\Z)', 'Hardware'),
            (r'Network.*?(?=\n[A-Z][A-Za-z\s]+:|\nInstallation|\nCommissioning|\Z)', 'Network'),
            (r'Installation.*?(?=\n[A-Z][A-Za-z\s]+:|\nCommissioning|\nValidation|\Z)', 'Installation'),
            (r'Commissioning.*?(?=\n[A-Z][A-Za-z\s]+:|\nValidation|\Z)', 'Commissioning'),
            (r'Validation.*?(?=\n[A-Z][A-Za-z\s]+:|\Z)', 'Validation')
        ]
        
        for pattern, section_name in section_patterns:
            match = re.search(pattern, full_text, re.DOTALL | re.IGNORECASE)
            if match:
                sections[section_name] = match.group(0).strip()
        
        return sections
    
    def _extract_site_survey_metadata(self, content: Dict[str, Any]) -> Dict[str, Any]:
        """Extract metadata from site survey content"""
        metadata = {}
        
        # Extract version information
        if 'Release Notes' in content:
            release_info = content['Release Notes'].get('release_info', {})
            if 'version' in release_info:
                metadata['template_version'] = release_info['version']
        
        # Extract project information
        if 'Project Details' in content:
            project_details = content['Project Details'].get('key_values', {})
            metadata.update({
                'customer_name': project_details.get('customer_name'),
                'project_name': project_details.get('project_name'),
                'opportunity_id': project_details.get('opportunity_id')
            })
        
        # Extract cluster information
        if 'VAST Cluster' in content:
            cluster_info = content['VAST Cluster'].get('cluster_info', {})
            metadata['cluster_description'] = cluster_info.get('cluster_description')
        
        return metadata
    
    def _extract_install_plan_metadata(self, full_text: str, sections: Dict[str, str]) -> Dict[str, Any]:
        """Extract metadata from install plan content"""
        metadata = {}
        
        # Extract project name
        project_match = re.search(r'Project:\s*(.+)', full_text, re.IGNORECASE)
        if project_match:
            metadata['project_name'] = project_match.group(1).strip()
        
        # Extract customer name
        customer_match = re.search(r'Customer:\s*(.+)', full_text, re.IGNORECASE)
        if customer_match:
            metadata['customer_name'] = customer_match.group(1).strip()
        
        # Extract version information
        version_match = re.search(r'VastOS\s+(\d+\.\d+\.\d+[-\d]*)', full_text)
        if version_match:
            metadata['vastos_version'] = version_match.group(1)
        
        # Extract cluster information
        cluster_match = re.search(r'(\d+x\d+)\s+cluster', full_text, re.IGNORECASE)
        if cluster_match:
            metadata['cluster_size'] = cluster_match.group(1)
        
        # Extract network information
        if 'Network' in sections:
            network_text = sections['Network']
            
            # Extract IP ranges
            ip_ranges = re.findall(r'\b(?:[0-9]{1,3}\.){3}[0-9]{1,3}(?:/\d+)?\b', network_text)
            if ip_ranges:
                metadata['ip_ranges'] = ip_ranges
            
            # Extract VLAN information
            vlan_matches = re.findall(r'VLAN\s*(\d+)', network_text, re.IGNORECASE)
            if vlan_matches:
                metadata['vlans'] = vlan_matches
        
        metadata['sections_found'] = list(sections.keys())
        
        return metadata

class GoogleSheetsProcessor:
    """Process Google Sheets documents"""
    
    def __init__(self, sheets_service):
        self.sheets_service = sheets_service
        self.logger = logging.getLogger(__name__)
    
    def process_google_sheet(self, sheet_id: str, sheet_range: str = None) -> ProcessedDocument:
        """Process a Google Sheets document"""
        try:
            # Get sheet metadata
            sheet_metadata = self.sheets_service.spreadsheets().get(
                spreadsheetId=sheet_id
            ).execute()
            
            # Get sheet data
            if sheet_range:
                result = self.sheets_service.spreadsheets().values().get(
                    spreadsheetId=sheet_id,
                    range=sheet_range
                ).execute()
                values = result.get('values', [])
            else:
                # Get all sheets
                values = {}
                for sheet in sheet_metadata.get('sheets', []):
                    sheet_name = sheet['properties']['title']
                    result = self.sheets_service.spreadsheets().values().get(
                        spreadsheetId=sheet_id,
                        range=sheet_name
                    ).execute()
                    values[sheet_name] = result.get('values', [])
            
            return ProcessedDocument(
                document_type='google_sheets',
                source_url=f"https://docs.google.com/spreadsheets/d/{sheet_id}",
                content={'values': values},
                metadata={
                    'sheet_id': sheet_id,
                    'title': sheet_metadata.get('properties', {}).get('title', ''),
                    'sheets': [s['properties']['title'] for s in sheet_metadata.get('sheets', [])]
                },
                processing_errors=[]
            )
            
        except Exception as e:
            self.logger.error(f"Error processing Google Sheet {sheet_id}: {str(e)}")
            return ProcessedDocument(
                document_type='google_sheets',
                source_url=f"https://docs.google.com/spreadsheets/d/{sheet_id}",
                content={},
                metadata={'error': str(e)},
                processing_errors=[str(e)]
            )

