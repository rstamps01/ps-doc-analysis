from flask import Blueprint, request, jsonify, current_app
import os
import tempfile
import json
from datetime import datetime
from werkzeug.utils import secure_filename
import PyPDF2
import openpyxl
from io import BytesIO

# Fix imports for standalone execution
import sys
sys.path.append(os.path.dirname(os.path.dirname(__file__)))
from integrations.google_drive import GoogleDriveIntegration
from integrations.google_sheets import GoogleSheetsIntegration

document_processing = Blueprint('document_processing', __name__)

ALLOWED_EXTENSIONS = {'pdf', 'xlsx', 'docx'}

def allowed_file(filename):
    return '.' in filename and \
           filename.rsplit('.', 1)[1].lower() in ALLOWED_EXTENSIONS

def extract_pdf_content(file_stream):
    """Extract text content from PDF file"""
    try:
        pdf_reader = PyPDF2.PdfReader(file_stream)
        content = ""
        for page in pdf_reader.pages:
            content += page.extract_text() + "\n"
        return content
    except Exception as e:
        return f"Error reading PDF: {str(e)}"

def extract_xlsx_content(file_stream):
    """Extract content from Excel file"""
    try:
        workbook = openpyxl.load_workbook(file_stream)
        content = {}
        for sheet_name in workbook.sheetnames:
            sheet = workbook[sheet_name]
            sheet_data = []
            for row in sheet.iter_rows(values_only=True):
                if any(cell is not None for cell in row):
                    sheet_data.append([str(cell) if cell is not None else "" for cell in row])
            content[sheet_name] = sheet_data
        return content
    except Exception as e:
        return f"Error reading Excel file: {str(e)}"

def analyze_document_content(content, filename):
    """Analyze document content against validation criteria"""
    
    # Mock validation logic - in real implementation, this would use the enhanced validation engine
    validation_results = {
        'total_criteria': 65,
        'categories': [
            {'name': 'Basic Project Information', 'total': 6, 'passed': 0, 'issues': []},
            {'name': 'SFDC & Documentation Integration', 'total': 2, 'passed': 0, 'issues': []},
            {'name': 'Template & Documentation Standards', 'total': 2, 'passed': 0, 'issues': []},
            {'name': 'Installation Plan Content Validation', 'total': 6, 'passed': 0, 'issues': []},
            {'name': 'Network Configuration & Technical', 'total': 9, 'passed': 0, 'issues': []},
            {'name': 'Site Survey Documentation', 'total': 12, 'passed': 0, 'issues': []},
            {'name': 'Cross-Document Consistency', 'total': 15, 'passed': 0, 'issues': []},
            {'name': 'Enhanced Features', 'total': 13, 'passed': 0, 'issues': []}
        ],
        'issues': [],
        'recommendations': []
    }
    
    # Basic content analysis
    if isinstance(content, str):  # PDF content
        content_lower = content.lower()
        
        # Basic Project Information checks
        if 'project' in content_lower and 'name' in content_lower:
            validation_results['categories'][0]['passed'] += 1
        else:
            validation_results['categories'][0]['issues'].append('Project name not clearly identified')
            
        if 'opportunity' in content_lower:
            validation_results['categories'][0]['passed'] += 1
        else:
            validation_results['categories'][0]['issues'].append('Opportunity ID missing')
            
        if 'customer' in content_lower:
            validation_results['categories'][0]['passed'] += 1
        else:
            validation_results['categories'][0]['issues'].append('Customer information incomplete')
            
        # Network Configuration checks
        if 'vlan' in content_lower:
            validation_results['categories'][4]['passed'] += 2
        else:
            validation_results['categories'][4]['issues'].append('VLAN configuration not documented')
            
        if 'ip address' in content_lower or 'network' in content_lower:
            validation_results['categories'][4]['passed'] += 2
        else:
            validation_results['categories'][4]['issues'].append('IP addressing scheme incomplete')
            
        if 'switch' in content_lower:
            validation_results['categories'][4]['passed'] += 1
        else:
            validation_results['categories'][4]['issues'].append('Switch configuration missing')
            
    elif isinstance(content, dict):  # Excel content
        sheet_names = list(content.keys())
        
        # Site Survey specific checks
        if any('project' in name.lower() for name in sheet_names):
            validation_results['categories'][0]['passed'] += 2
        else:
            validation_results['categories'][0]['issues'].append('Project Details worksheet missing')
            
        if any('network' in name.lower() for name in sheet_names):
            validation_results['categories'][4]['passed'] += 3
            validation_results['categories'][5]['passed'] += 2
        else:
            validation_results['categories'][4]['issues'].append('Network Details worksheet missing')
            validation_results['categories'][5]['issues'].append('Network documentation incomplete')
            
        if any('rack' in name.lower() for name in sheet_names):
            validation_results['categories'][5]['passed'] += 3
        else:
            validation_results['categories'][5]['issues'].append('Rack diagram missing')
            
        if any('hardware' in name.lower() for name in sheet_names):
            validation_results['categories'][5]['passed'] += 2
        else:
            validation_results['categories'][5]['issues'].append('Hardware details incomplete')
    
    # Calculate overall score
    total_passed = sum(cat['passed'] for cat in validation_results['categories'])
    score = total_passed / validation_results['total_criteria']
    
    # Determine status
    status = 'passed' if score >= 0.8 else 'partial' if score >= 0.6 else 'failed'
    
    # Collect all issues
    all_issues = []
    for category in validation_results['categories']:
        all_issues.extend(category['issues'])
    
    validation_results['score'] = round(score, 3)
    validation_results['status'] = status
    validation_results['issues'] = all_issues
    validation_results['passed_criteria'] = total_passed
    
    return validation_results

@document_processing.route('/api/documents/upload', methods=['POST'])
def upload_document():
    """Handle document upload"""
    try:
        if 'file' not in request.files:
            return jsonify({'error': 'No file provided'}), 400
        
        file = request.files['file']
        if file.filename == '':
            return jsonify({'error': 'No file selected'}), 400
        
        if not allowed_file(file.filename):
            return jsonify({'error': 'File type not allowed. Please upload PDF, XLSX, or DOCX files.'}), 400
        
        filename = secure_filename(file.filename)
        file_id = f"{datetime.now().strftime('%Y%m%d_%H%M%S')}_{filename}"
        
        # Save file temporarily for processing
        temp_dir = tempfile.gettempdir()
        file_path = os.path.join(temp_dir, file_id)
        file.save(file_path)
        
        return jsonify({
            'success': True,
            'file_id': file_id,
            'filename': filename,
            'size': os.path.getsize(file_path),
            'upload_time': datetime.now().isoformat()
        })
        
    except Exception as e:
        return jsonify({'error': f'Upload failed: {str(e)}'}), 500

@document_processing.route('/api/documents/process-google-drive', methods=['POST'])
def process_google_drive_url():
    """Process a Google Drive URL and download the file"""
    try:
        data = request.get_json()
        if not data or 'url' not in data:
            return jsonify({'error': 'Google Drive URL is required'}), 400
        
        url = data['url'].strip()
        if not url:
            return jsonify({'error': 'Google Drive URL cannot be empty'}), 400
        
        # Initialize Google Drive integration
        drive_integration = GoogleDriveIntegration()
        if not drive_integration.service:
            return jsonify({'error': 'Google Drive integration not configured'}), 500
        
        # Process the Google Drive URL
        result = drive_integration.process_google_drive_url(url)
        
        if result['success']:
            # Generate file ID for our system
            file_id = f"gdrive_{datetime.now().strftime('%Y%m%d_%H%M%S')}_{result['file_id']}"
            
            return jsonify({
                'success': True,
                'file_id': file_id,
                'filename': result['metadata'].get('file_name', 'Unknown'),
                'size': result['metadata'].get('size', 'Unknown'),
                'upload_time': datetime.now().isoformat(),
                'source': 'google_drive',
                'original_url': url,
                'local_path': result['local_path'],
                'metadata': result['metadata']
            })
        else:
            return jsonify({
                'success': False,
                'error': result['error'],
                'url': url
            }), 400
            
    except Exception as e:
        return jsonify({'error': f'Google Drive processing failed: {str(e)}'}), 500

@document_processing.route('/api/documents/process-google-sheets', methods=['POST'])
def process_google_sheets_url():
    """Process a Google Sheets URL and extract validation criteria"""
    try:
        data = request.get_json()
        if not data or 'url' not in data:
            return jsonify({'error': 'Google Sheets URL is required'}), 400
        
        url = data['url'].strip()
        if not url:
            return jsonify({'error': 'Google Sheets URL cannot be empty'}), 400
        
        # Extract spreadsheet ID from URL
        import re
        match = re.search(r'/spreadsheets/d/([a-zA-Z0-9-_]+)', url)
        if not match:
            return jsonify({'error': 'Invalid Google Sheets URL format'}), 400
        
        spreadsheet_id = match.group(1)
        
        # Initialize Google Sheets integration
        sheets_integration = GoogleSheetsIntegration()
        if not sheets_integration.service:
            return jsonify({'error': 'Google Sheets integration not configured'}), 500
        
        # Get sheet metadata
        try:
            metadata = sheets_integration.get_sheet_metadata(spreadsheet_id)
        except Exception as e:
            return jsonify({'error': f'Cannot access Google Sheets document: {str(e)}'}), 400
        
        # Read requirements from the sheet
        try:
            requirements = sheets_integration.read_requirements(spreadsheet_id)
        except Exception as e:
            return jsonify({'error': f'Error reading sheet data: {str(e)}'}), 400
        
        # Generate file ID for our system
        file_id = f"gsheets_{datetime.now().strftime('%Y%m%d_%H%M%S')}_{spreadsheet_id}"
        
        return jsonify({
            'success': True,
            'file_id': file_id,
            'filename': metadata.get('title', 'Google Sheets Document'),
            'upload_time': datetime.now().isoformat(),
            'source': 'google_sheets',
            'original_url': url,
            'spreadsheet_id': spreadsheet_id,
            'metadata': metadata,
            'requirements_count': len(requirements),
            'requirements': requirements[:10] if len(requirements) > 10 else requirements  # Limit preview
        })
        
    except Exception as e:
        return jsonify({'error': f'Google Sheets processing failed: {str(e)}'}), 500

@document_processing.route('/api/documents/validate/<file_id>', methods=['POST'])
def validate_document(file_id):
    """Validate uploaded document"""
    try:
        # Determine file source and path
        if file_id.startswith('gdrive_'):
            # Google Drive file - need to find the downloaded file
            temp_dir = '/tmp/validation_downloads'
            # Extract original file ID and find the file
            parts = file_id.split('_', 3)
            if len(parts) >= 4:
                original_file_id = parts[3]
                # Look for files in download directory
                if os.path.exists(temp_dir):
                    for filename in os.listdir(temp_dir):
                        if original_file_id in filename or file_id in filename:
                            file_path = os.path.join(temp_dir, filename)
                            break
                    else:
                        return jsonify({'error': 'Downloaded file not found'}), 404
                else:
                    return jsonify({'error': 'Download directory not found'}), 404
            else:
                return jsonify({'error': 'Invalid Google Drive file ID format'}), 400
        elif file_id.startswith('gsheets_'):
            # Google Sheets file - handle validation
            parts = file_id.split('_', 3)
            if len(parts) >= 4:
                spreadsheet_id = parts[3]
                
                # Import the Google Sheets validator
                from validation.google_sheets_validator import GoogleSheetsValidator
                
                # Initialize validator
                sheets_validator = GoogleSheetsValidator()
                
                # Determine document type and validate accordingly
                # For now, assume it's validation criteria - in production, this would be determined from metadata
                validation_results = sheets_validator.validate_validation_criteria_sheet(spreadsheet_id)
                
                return jsonify({
                    'success': True,
                    'validation_results': validation_results
                })
            else:
                return jsonify({'error': 'Invalid Google Sheets file ID format'}), 400
        else:
            # Regular uploaded file
            temp_dir = tempfile.gettempdir()
            file_path = os.path.join(temp_dir, file_id)
        
        if not os.path.exists(file_path):
            return jsonify({'error': 'File not found'}), 404
        
        # Extract content based on file type
        filename = os.path.basename(file_path)
        file_ext = filename.rsplit('.', 1)[1].lower() if '.' in filename else ''
        
        with open(file_path, 'rb') as f:
            if file_ext == 'pdf':
                content = extract_pdf_content(f)
            elif file_ext == 'xlsx':
                content = extract_xlsx_content(f)
            else:
                return jsonify({'error': 'Unsupported file type for validation'}), 400
        
        # Analyze content
        validation_results = analyze_document_content(content, filename)
        
        # Add metadata
        validation_results.update({
            'file_id': file_id,
            'filename': filename,
            'processed_time': datetime.now().isoformat(),
            'file_type': file_ext,
            'source': 'google_drive' if file_id.startswith('gdrive_') else 'upload'
        })
        
        # Clean up temporary file for regular uploads only
        if not file_id.startswith('gdrive_'):
            try:
                os.remove(file_path)
            except:
                pass  # Ignore cleanup errors
        
        return jsonify({
            'success': True,
            'validation_results': validation_results
        })
        
    except Exception as e:
        return jsonify({'error': f'Validation failed: {str(e)}'}), 500

@document_processing.route('/api/documents/results', methods=['GET'])
def get_validation_results():
    """Get all validation results"""
    # In a real implementation, this would fetch from database
    # For now, return mock data
    return jsonify({
        'success': True,
        'results': []
    })

@document_processing.route('/api/documents/criteria', methods=['GET'])
def get_validation_criteria():
    """Get validation criteria information"""
    criteria = {
        'total_criteria': 65,
        'categories': [
            {
                'name': 'Basic Project Information',
                'count': 6,
                'description': 'Project name, opportunity ID, customer details, timeline, approvals',
                'checks': [
                    'Project name clearly identified',
                    'Opportunity ID present and valid',
                    'Customer information complete',
                    'Project timeline defined',
                    'Required approvals obtained',
                    'Contact information current'
                ]
            },
            {
                'name': 'SFDC & Documentation Integration',
                'count': 2,
                'description': 'Salesforce integration and documentation links',
                'checks': [
                    'SFDC opportunity link valid',
                    'Documentation references accurate'
                ]
            },
            {
                'name': 'Template & Documentation Standards',
                'count': 2,
                'description': 'Template compliance and documentation standards',
                'checks': [
                    'Template version current',
                    'Documentation standards followed'
                ]
            },
            {
                'name': 'Installation Plan Content Validation',
                'count': 6,
                'description': 'Installation procedures and technical specifications',
                'checks': [
                    'Installation procedures documented',
                    'Technical specifications complete',
                    'Prerequisites identified',
                    'Risk assessment included',
                    'Rollback procedures defined',
                    'Testing procedures outlined'
                ]
            },
            {
                'name': 'Network Configuration & Technical',
                'count': 9,
                'description': 'Network setup, VLAN configuration, IP addressing',
                'checks': [
                    'VLAN configuration documented',
                    'IP addressing scheme defined',
                    'Switch configuration specified',
                    'Network topology clear',
                    'Security settings documented',
                    'Bandwidth requirements specified',
                    'Network validation procedures',
                    'Firewall configuration',
                    'DNS/DHCP settings'
                ]
            },
            {
                'name': 'Site Survey Documentation',
                'count': 12,
                'description': 'Physical site requirements and constraints',
                'checks': [
                    'Rack space requirements',
                    'Power requirements documented',
                    'Cooling requirements specified',
                    'Physical access documented',
                    'Environmental conditions',
                    'Cable routing planned',
                    'Hardware inventory complete',
                    'Site contact information',
                    'Delivery logistics',
                    'Installation timeline',
                    'Site-specific constraints',
                    'Safety requirements'
                ]
            },
            {
                'name': 'Cross-Document Consistency',
                'count': 15,
                'description': 'Consistency between Site Survey parts and Install Plan',
                'checks': [
                    'Hardware specs consistent',
                    'Network config aligned',
                    'Timeline synchronization',
                    'Contact info matches',
                    'Version consistency',
                    'Approval status aligned',
                    'Technical requirements match',
                    'Site details consistent',
                    'Project scope aligned',
                    'Resource allocation consistent',
                    'Risk assessments aligned',
                    'Testing procedures match',
                    'Documentation references consistent',
                    'Change management aligned',
                    'Quality assurance consistent'
                ]
            },
            {
                'name': 'Enhanced Features',
                'count': 13,
                'description': 'Advanced validation capabilities',
                'checks': [
                    'Conditional logic processing',
                    'Automation complexity classification',
                    'Confidence scoring',
                    'Real-time accuracy monitoring',
                    'Pattern recognition',
                    'Content analysis',
                    'Cross-reference validation',
                    'Intelligent prioritization',
                    'Adaptive validation rules',
                    'Machine learning insights',
                    'Predictive analysis',
                    'Quality trend analysis',
                    'Continuous improvement tracking'
                ]
            }
        ],
        'enhanced_features': [
            'Conditional Logic Processing',
            'Cross-Document Validation',
            'Automation Complexity Classification',
            'Confidence Scoring',
            'Real-Time Accuracy Monitoring'
        ]
    }
    
    return jsonify({
        'success': True,
        'criteria': criteria
    })

